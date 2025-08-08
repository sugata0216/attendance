from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Attendance, Student, Subject
import datetime
from django.views import generic
from .forms import AttendanceForm
from django.contrib.auth.decorators import login_required
import csv
from django.db.models import Count
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
def index(request):
    return render(request, 'attendance_management/index.html')
def help(request):
    return render(request, 'attendance_management/help.html')
def create_attendance(request, date):
    if request.method == 'POST':
        form = AttendanceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('attendance_management:index')
    else:
        form = AttendanceForm(initial={'date' : date})
    return render(request, 'attendance_management/create.html', {
        'form':form,
        'date':date,
        'object':None
        })
def update_attendance(request, date, time_limit):
    attendance = get_object_or_404(Attendance, date=date, time_limit=time_limit, student=request.user.student)
    if request.method == 'POST':
        form = AttendanceForm(request.POST, instance=attendance)
        if form.is_valid():
            form.save()
            return redirect('attendance_management:index')
    else:
        form = AttendanceForm(instance=attendance)
    return render(request, 'attendance_management/create.html', {
        'form':form,
        'date':date,
        'object':attendance
    })
@login_required
def create_or_update_attendance(request, date, time_limit):
    time_limit = int(time_limit)
    try:
        attendance = Attendance.objects.get(date=date, time_limit=time_limit, student=request.user.student)
        is_update = True
    except Attendance.DoesNotExist:
        attendance = None
        is_update = False
    if request.method == 'POST':
        form = AttendanceForm(request.POST, instance=attendance)
        if form.is_valid():
            new_attendance = form.save(commit=False)
            new_attendance.date = date
            new_attendance.time_limit = time_limit
            new_attendance.student = request.user.student
            new_attendance.save()
            return redirect('attendance_management:index')
    else:
        form = AttendanceForm(instance=attendance)
    return render(request, 'attendance_management/create.html', {
        'form' : form,
        'attendance' : attendance,
        'date' : date,
        'time_limit' : time_limit,
        'is_update' : is_update,
    })
@csrf_exempt
def mark_attendance(request):
    if request.method == 'POST':
        # POSTデータを取得
        date = request.POST.get('date')
        student_id = request.POST.get('student_id')
        subject_id = request.POST.get('subject_id')
        status = request.POST.get('status')

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Invalid method'}, status=400)
def get_events(request):
    if request.method == 'GET':
        events = []
        for attendance in Attendance.objects.all():
            events.append({
                'title': f"{attendance.subject.name}（{attendance.student.name}）",
                'start': attendance.date.strftime("%Y-%m-%dT%H:%M:%S"),
                'end': attendance.date.strftime("%Y-%m-%dT%H:%M:%S"),
            })
        return JsonResponse(events, safe=False)
def attendance_delete(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)
    if request.method == "POST":
        attendance.delete()
        return redirect('attendance_management:index')
    return render(request, 'attendance_management/delete.html', {'attendance' : attendance})
@login_required
def export_attendance_csv(request):
    student = request.user.student
    responce = HttpResponse(content_type='text/csv')
    responce['Content-Disposition'] = 'attachment; filename="attendance.csv"'
    writer = csv.writer(responce)
    writer.writerow(['日付', '時限', '出席状況'])
    attendance_records = Attendance.objects.filter(student=student).order_by('date', 'time_limit')
    for record in attendance_records:
        writer.writerow([
            record.date.strftime('%Y-%m-%d'),
            record.time_limit,
            record.get_status_display()
        ])
    return responce
@login_required
def attendance_summary(request):
    student = request.user.student
    total = Attendance.objects.filter(student=student).count()
    present = Attendance.objects.filter(student=student, status='present').count()
    absent = Attendance.objects.filter(student=student, status='absent').count()
    late = Attendance.objects.filter(student=student, status='late').count()
    leave_early = Attendance.objects.filter(student=student, status='leave_early').count()
    data = {
        'total': total,
        'present': present,
        'absent': absent,
        'late': late,
        'early': leave_early,
        'student_name': student.name,
    }
    return render(request, 'attendance_management/attendance_summary.html', data)
@login_required
def attendance_by_subject(request):
    student = request.user.student
    subjects = Subject.objects.all()
    data = []
    for subject in subjects:
        total = Attendance.objects.filter(student=student, subject=subject).count()
        present = Attendance.objects.filter(student=student, subject=subject, status='present').count()
        absent = Attendance.objects.filter(student=student, subject=subject, status='absent').count()
        late = Attendance.objects.filter(student=student, subject=subject, status='late').count()
        leave_early = Attendance.objects.filter(student=student, subject=subject, status='leave_early').count()
        if total > 0:
            data.append({
                'subject_name': subject.name,
                'total': total,
                'present': present,
                'absent': absent,
                'late': late,
                'leave_early': leave_early,
            })
    context = {
        'data': data,
    }
    return render(request, 'attendance_management/attendance_by_subject.html', context)
@login_required
def download_attendance_pdf(request):
    student = request.user.student
    attendances = Attendance.objects.filter(student=student).order_by('-date')
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
    p.setFont('HeiseiKakuGo-W5', 16)
    p.drawString(100, height - 50, f"{student.name}さんの出席履歴")
    p.setFont("HeiseiKakuGo-W5", 12)
    y = height - 80
    for attendance in attendances:
        line = f"{attendance.date} - {attendance.get_time_limit_display()} - {attendance.subject.name} - {attendance.get_status_display()}"
        if attendance.reason:
            line += f"(理由:{attendance.reason})"
        p.drawString(50, y, line)
        y -= 20
        if y < 50:
            p.showPage()
            p.setFont("HeiseiKakuGo-W5", 12)
            y = height - 50
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')
@login_required
def attendance_history(request):
    student = request.user.student
    attendance_records = Attendance.objects.filter(student=student).order_by('-date', 'time_limit')
    return render(request, 'attendance_management/attendance_history.html', {
        'attendance_records': attendance_records
    })
@login_required
def export_attendance_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出席履歴"
    ws.append(["日付", "授業名", "限目", "出席状況", "理由"])
    student = Student.objects.get(user=request.user)
    attendances = Attendance.objects.filter(student=student).order_by('-date')
    for a in attendances:
        ws.append([a.date, a.subject.name, f"{a.time_limit}限目", a.get_status_display(), a.reason or ""])
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Discription'] = 'attachment; filename=attendance.xlsx'
    wb.save(response)
    return response
# Create your views here.